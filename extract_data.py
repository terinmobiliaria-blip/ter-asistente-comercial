from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\MKT 3\Desktop\TER_ArgumentosComerciales_v3.xlsx")
OUTPUT = Path(__file__).with_name("app-data.js")


def clean_project(value: object, fallback: str) -> str:
    text = str(value or fallback).strip()
    text = re.sub(r"^[^A-Za-zÁÉÍÓÚÑáéíóúñ0-9]+\s*", "", text).strip()
    return text


def main() -> None:
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    excluded = {
        "📖 Cómo usar este documento",
        "🎨 Leyenda de Colores",
    }
    rows = []

    for ws in wb.worksheets:
        if ws.title in excluded:
            continue

        is_finish_sheet = ws.title == "✨ Cuadro de Acabados"
        project = "Acabados generales" if is_finish_sheet else clean_project(ws["A1"].value, ws.title)

        for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
            feature, category, advantage, benefit, speech, profile, objection = row
            if not feature or not category or "▸" in str(feature):
                continue

            rows.append(
                {
                    "kind": "finish" if is_finish_sheet else "project",
                    "project": project,
                    "category": str(category).strip(),
                    "feature": str(feature).strip(),
                    "advantage": str(advantage or "").strip(),
                    "benefit": str(benefit or "").strip(),
                    "speech": str(speech or "").strip(),
                    "profile": str(profile or "Todos").strip(),
                    "objection": str(objection or "").strip(),
                }
            )

    OUTPUT.write_text(
        "window.TER_ARGUMENTOS = "
        + json.dumps(rows, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Argumentos exportados: {len(rows)}")
    print(OUTPUT)


if __name__ == "__main__":
    main()
